from openpyxl import *
import os
from Config import *
class rewrxl():
    def __init__(self):
        self.path = path_xlexcel()
        # print(self.path)

    def write_xl(self,col,value):
        '''自增行写入学员数据'''
        wb = load_workbook(self.path)
        #打开sheet
        ws = wb["ApiStudentBaseInfo"]
        #方法1
        # ws.cell(row=3, column=3).value = 'test'
        #方法2
        row=ws.max_row
        if col=='A':
            ws['%s%d' %(col,row+1)]= value
        else:
            ws['%s%d' % (col, row)] = value
        wb.save(self.path)

    def sread_xl(self,SheetName,col):
        '''读取接口所需数据'''
        wb = load_workbook(self.path)
        ws = wb.get_sheet_by_name(SheetName)
        #ws = wb.get_sheet_by_name("StudentNum")
        row=ws.max_row
        sheet_value=ws['%s%d' % (col, row)].value
        return sheet_value


    def swrite_xl(self,SheetName,col,value):
        '''自增行写入学员接口数据数据'''
        wb = load_workbook(self.path)
        ws = wb[SheetName]
        # 方法1
        # ws.cell(row=3, column=3).value = 'test'
        # 方法2
        if SheetName == 'StudentNum':
            ws['%s2' %col] = value
            wb.save(self.path)
        elif SheetName == 'Cookies':
            ws['%s1' %col] = value
            wb.save(self.path)
        else:
            row = ws.max_row
            if col == 'A':
                ws['%s%d' % (col, row + 1)] = value
            else:
                ws['%s%d' % (col, row)] = value
            wb.save(self.path)


if __name__ == '__main__':
    l=rewrxl()
    l.swrite_xl('Cookies','D', 'fafasgafasf')
